#!/usr/local/bin/python
# -*- coding: utf-8 -*-
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.formatting import Rule
# from openpyxl.styles.differential import DifferentialStyle
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.pagebreak import Break  # , RowBreak, ColBreak
import tkinter
from tkinter import filedialog as fd
# import re
import os
import sys

#   TODO global variable
temp = ''
filetypes = (('text files', '*.txt'), ('All files', '*.*'))
device_type = {
    1: ["Сигнал-20"],
    2: ["Сигнал-20П"],
    3: ["С2000-СП1"],
    4: ["С2000-4"],
    7: ["С2000-К"],
    8: ["С2000-ИТ"],
    9: ["С2000-КДЛ"],
    10: ["С2000-БИ/БКИ"],
    11: ["Сигнал-20(вер. 02)"],
    12: [],
    13: ["С2000-КС"],
    14: ["С2000-АСПТ"],
    15: ["С2000-КПБ"],
    16: ["С2000-2"],
    17: [],
    18: [],
    19: ["УО-ОРИОН"],
    20: ["Рупор"],
    21: ["Рупор-Диспетчер исп.01"],
    22: ["С2000-ПТ"],
    23: [],
    24: ["УО-4С"],
    25: ["Поток-3Н"],
    26: ["Сигнал-20М"],
    27: [],
    28: ["С2000-БИ-01"],
    29: [],
    30: ["Рупор-01"],
    31: ["С2000-Adem"],
    32: ["РИП-12 исп.50, исп.51, без исполнения"],
    33: ["РИП-12 исп.50, исп.51, без исполнения"],
    34: ["Сигнал-10"],
    35: ["РИП-12 исп.54"],
    36: ["С2000-ПП"],
    37: ["РИП-24 исп.50, исп.51"],
    38: ["РИП-12 исп.54"],
    39: ["РИП-24 исп.50, исп.51"],
    40: [],
    41: ["С2000-КДЛ-2И"],
    42: [],
    43: ["С2000-PGE"],
    44: ["С2000-БКИ"],
    45: ["Поток-БКИ"],
    46: ["Рупор-200"],
    47: ["С2000-Периметр"],
    48: ["МИП-12"],
    49: ["МИП-24"],
    50: [],
    51: [],
    52: [],
    53: ["РИП-48 исп.01"],
    54: ["РИП-12 исп.56"],
    55: ["РИП-24 исп.56"],
    56: [],
    57: [],
    58: [],
    59: ["Рупор исп.02"],
    60: [],
    61: ["С2000-КДЛ-Modbus"],
    66: ["Рупор исп.03"],
    67: ["Рупор-300"],
    76: ["С2000-PGE исп.01"]
}  # TODO типы приборов
cable_type = {
    0: ['по умолчанию'],
    1: ['охранный'],
    2: ['пожарный'],
    3: ['тревожный'],
    4: ['технологический'],
    5: ['входной'],
    6: ['адресно-аналоговый дымовой'],
    7: ['адресно-аналоговый тепловой'],
    8: ['8 тип - хз'],
    9: ['цепь ДС дверей'],
    10: ['ручной пуск'],
    11: ['дистанционный пуск'],
    12: ['состояние автоматики']
}  # TODO типы шлейфов
relay = {
    1: ['включить'],
    2: ['выключить'],
    3: ['включить на время'],
    4: ['выключить на время'],
    5: ['мигать из состояния выключено'],
    6: ['мигать из состояния включено'],
    7: ['мигать из состояния выключено'],
    8: ['мигать из состояния включено'],
    9: ['лампа'],
    10: ['пцн'],
    11: ['аспт'],
    12: ['сирена'],
    13: ['пожарный пцн'],
    14: ['выход неисправность'],
    15: ['пожарная лампа'],
    16: ['старая тактика пцн'],
    17: ['включить на время перед взятием'],
    18: ['выключить на время перед взятием'],
    19: ['включить на время при взятии'],
    20: ['выключить на время при взятии'],
    21: ['включить на время при снятии'],
    22: ['включить на время при снятии'],
    23: ['включить на время при невзятии'],
    24: ['выключить на время при невзятии'],
    25: ['включить на время при нарушении'],
    26: ['включить на время при нарушении'],
    27: ['включить при снятии'],
    28: ['выключить при снятии'],
    29: ['включить при взятии'],
    30: ['включить при взятии'],
    31: ['включить при нарушении тех. шлейфа'],
    32: ['вsключить при нарушении тех. шлейфа'],
    33: ['аспт-1'],
    34: ['аспт-а'],
    35: ['аспт-а1'],
    36: ['включить при повышении температуры'],
    37: ['выключить при повышении температуры'],
    38: ['включить при задержке пуска'],
    39: ['включить при пуске'],
    40: ['включить при тушении'],
    41: ['включить при неудачном пуске'],
    42: ['включить при включении автоматики'],
    43: ['выключить при включении автоматики'],
    44: ['включить при выключении автоматики'],
    45: ['выключить при выключении автоматики'],
    46: ['вкл. если исп. устройство в рабочем состоянии'],
    47: ['вык. если исп. устройство в рабочем состоянии'],
    48: ['вкл. если исп. устройство в исходном состоянии'],
    49: ['вык. если исп. устройство в исходном состоянии'],
    50: ['включить при пожар2'],
    51: ['выключить при пожар2'],
    52: ['мигать при пожар2; исх. сост. выключено'],
    53: ['мигать при пожар2; исх. сост. включено'],
    54: ['включить при нападении'],
    55: ['выключить при нападении'],
    56: ['лампа 2'],
    57: ['сирена 2']
}  # TODO реле
cable_script = {
    1: ['Нет'],
    2: ['Снять шлейф'],
    3: ['Взять шлейф'],
    4: ['Сбросить тревогу'],
    5: ['Откл. Автоматику'],
    6: ['Вкл. Автоматику'],
    7: ['Отменить пуск АУП'],
    8: ['Запустить АУП'],
    9: ['Вкл. режим тестирования'],
    10: ['Откл. режим тестирования']
}  # TODO сценарии, управление шлейфом
relay_script = {
    1: ['включить'],
    2: ['выключить'],
    3: ['включить на время'],
    4: ['выключить на время'],
    5: ['мигать из состояния выключено'],
    6: ['мигать из состояния включено'],
    7: ['мигать из состояния выключено на время'],
    8: ['мигать из состояния включено на время'],
    9: ['лампа'],
    10: ['пцн'],
    11: ['аспт'],
    12: ['нет']
}  # TODO сценарии, управление реле
mask = {
    1: ['****************'],
    2: ['----------------'],
    3: ['********--------'],
    4: ['--------********'],
    5: ['****----****----'],
    6: ['----****----****'],
    7: ['****------------'],
    8: ['----************'],
    9: ['**--**--**--**--'],
    10: ['--**--**--**--**'],
    11: ['**----**----**--'],
    12: ['--****--****--**'],
    13: ['**--------------'],
    14: ['--**************'],
    15: ['**--**__********'],
    16: ['--**--**--------'],
    17: ['*-*-*-*-*-*-*-*-'],
    18: ['-*-*-*-*-*-*-*-*'],
    19: ['*--*--*--*--*--*'],
    20: ['-**-**-**-**-**-'],
    21: ['*-------*-------'],
    22: ['-*******-*******'],
    23: ['*---------------'],
    24: ['-***************'],
    25: ['*-*-----*-*-----'],
    26: ['-*-*****-*-*****'],
    27: ['*-*-------------'],
    28: ['-*-*************'],
    29: ['*-*-*-----------'],
    30: ['-*-*-***********']
}  # TODO маска мигания
section = dict()  # TODO разделы

#   TODO создаем xlsx таблицу
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Адреса, шлейфа'


#   TODO инфа по сценариям (не дописано, в процессе)
def Exls_w_sc(row=1, column=1, list_str=None, bold=False, group_start=1):
    global temp
    level = 1
    for cel in list_str:
        if cel.find('Сценарий_упр:') != -1:
            column = 1
            ws.row_dimensions[row].height = 37
        elif cel.find('Описание:') != -1:
            column = 6
        elif cel.find('Тип:') != -1:
            temp = cel[cel.find('Тип: '):].replace('Тип: ', '').strip('""')
            print(temp)
            column = 2
        elif cel.find('Исх. программа упр.:') != -1:
            int_ish = int(cel.replace('Исх. программа упр.:', '').replace(' ', ''))
            cel = 'Исх. пр. упр: ' + str(relay_script[int_ish]).strip('[]')
            column = 3
        elif cel.find('Исх. маска мигания:') != -1:
            column = 4
        elif cel.find('Шаг:') != -1:
            ws.row_dimensions[row].height = 37
            column = 2
        elif cel.find('Программа:') != -1 and str(ws.cell(row, 2).value).find('Шаг:') != -1:
            int_sc = int(cel.replace(' ', '').replace('Программа:', ''))
            if temp == 'управление реле':
                cel = 'Пр. упр: ' + str(relay_script[int(int_sc)]).strip('[]').strip('\'\'')
            if temp == 'управление шлейфом':
                cel = 'Пр. упр: ' + str(cable_script[int(int_sc)]).strip('[]').strip('\'\'')
            ws.row_dimensions[row].height = 37
            column = 3
        elif cel.find('Состояния:') != -1:
            ws.row_dimensions[row].height = 37
            column = 4
            row -= 1
        elif cel.find('Разделы: ') != -1:
            #                print(cel.find('Разделы: '))
            cel = section[int(cel[int(cel.find('Разделы: ')):].replace('Разделы: ', ''))]
            cel = cel[cel.find('Описание: '):].replace('Описание: ', '')
            ws.row_dimensions[row].height = 25
            column = 6
        elif cel.find('Разрешающее условие') != -1:
            cel = cel[cel.find('Разрешающее условие'):]
            ws.row_dimensions[row].height = 37
            column = 3
        elif cel.find('Запрещающее условие') != -1:
            cel = cel[cel.find('Запрещающее условие'):]
            ws.row_dimensions[row].height = 37
            column = 3
        elif cel.find('Маска мигания:') != -1:
            print('Маска мигания:')
            continue
        elif cel.find('Задержка включения:') != -1:
            print('Задержка включения:')
            continue
        elif cel.find('Время управления:') != -1:
            print('Время управления:')
            continue
        elif cel.find('Статус:') != -1:
            print('Статус:')
            continue
        else:
            '''
            cel = type_r[int(cel)]
            cel = cel[cel.find('Описание: '):].replace('Описание: ', '')
            ws.row_dimensions[row].height = 25
            column = 6
            '''
            row += 1
        #                int_r = int(cel)
        cells = ws.cell(row=row, column=column, value=cel)
        cells.font = Font(name='Times New Roman', size=10, bold=bold, italic=True)
        cells.alignment = Alignment(wrap_text=True)
        #   TODO    если четная строка
        cells.fill = PatternFill(fill_type='solid', fgColor='F0F8FF') if row % 2 == 0 else PatternFill(
            fill_type='solid', fgColor='FFFAF0')
        column += 1
    #   TODO группируем по строкамм
    ws.row_dimensions.group(group_start, row, outline_level=level, hidden=False)
    row += 1
    return row


def exls_w_titul(row=1, column=1, list_str=None, bold=True):
    ws.merge_cells('A1:C1')
    for cel in list_str:
        if cel.find('Конфигурация ') != -1:
            column = 1
        if cel.find('Версия:') != -1:
            column = 4
        cells = ws.cell(row=row, column=column, value=cel)
        cells.alignment = Alignment(wrap_text=True)
        cells.font = Font(name='Times New Roman', size=10, bold=bold, italic=True)
        #   TODO    если четная строка
        cells.fill = PatternFill(fill_type='solid', fgColor='F0F8FF') if row % 2 == 0 else PatternFill(
            fill_type='solid', fgColor='FFFAF0')
        column += 1
    row += 1
    return row


def exls_w_out(row=1, column=1, list_str=None, bold=True):
    global section
    #    thins = Side(border_style="hair", color="000000")
    group_start = row
    level = 1
    for cel in list_str:
        #   TODO column 2
        if cel.find('Шлейф:') != -1:
            column = 2
            level = 1
        elif cel.find('Выход:') != -1:
            column = 2
            level = 2
        elif cel.find('Реле:') != -1:
            column = 2
            level = 3
            ws.row_dimensions[row].height = 25
        elif cel.find('Раздел:') != -1:
            int_r = int(cel.replace(' ', '').replace('Раздел:', ''))
            cel = str(section[int_r]).replace('Описание: ', '')[2:]
            ws.row_dimensions[row].height = 25
            column = 3
        elif cel.find('Программа:') != -1 and str(ws.cell(row, 2).value).find('Реле:') != -1:
            int_relay = cel.replace(' ', '').replace('Программа:', '')
            cel = 'Пр. упр: ' + str(relay[int(int_relay)]).strip('[]').strip('\'\'')
            if 20 < cel.__len__() < 30:
                ws.row_dimensions[row].height = 25
            if cel.__len__() > 30:
                ws.row_dimensions[row].height = 37
            column = 3
        elif cel.find('Тип_шлейфа:') != -1:
            int_cable = int(cel.replace(' ', '').replace('Тип_шлейфа:', ''))
            cel = str(cable_type[int_cable]).strip('[]')
            ws.row_dimensions[row].height = 25
            column = 4
        elif cel.find('Описание:') != -1:
            cel = cel.replace('Описание:', '')
            column = 6
        elif cel.find('Время') != -1:
            column = 4
        cells = ws.cell(row=row, column=column, value=cel)
        cells.font = Font(name='Times New Roman', size=10, bold=bold, italic=True)
        cells.alignment = Alignment(wrap_text=True)
        #   TODO    если четная строка
        cells.fill = PatternFill(fill_type='solid', fgColor='F0F8FF') if row % 2 == 0 else PatternFill(
            fill_type='solid', fgColor='FFFAF0')

        column += 1
    #   TODO группируем по строкамм
    ws.row_dimensions.group(start=group_start, end=row, outline_level=level, hidden=False)
    row += 1
    return row


def exls_w_adr(row=1, column=1, list_str=None, bold=True):
    title_sh = ['', 'Шлейф', 'Раздел', 'Тип шлейфа', '', 'Описание']
    thins = Side(border_style="hair", color="000000")
    #   TODO заполняем строку адреса
    for cel in list_str:
        if cel.find('Адрес:') != -1:
            column = 1
        if cel.find('Тип_прибора: ') != -1:
            cel = str(device_type[int(cel[cel.find('Тип_прибора: '):].replace('Тип_прибора: ', ''))]).strip('[]').strip(
                '\'\'')
        if column == 2 and cel.find('Сценарий_упр:') == -1:
            t_cell = ' '
            cells = ws.cell(row=row, column=3, value=t_cell)
            cells.border = Border(bottom=thins, top=thins, left=thins, right=thins)
            cells.fill = PatternFill(fill_type='solid', fgColor='E6E6FA')
        if cel.find('Сценарий_упр:') != -1:
            column = 3
        if cel.find('Версия:') != -1:
            column = 4
        if cel.find('Описание:') != -1:
            cel = cel.replace('Описание:', '').replace('\"', '')
            column = 6
        cells = ws.cell(row=row, column=column, value=cel)
        cells.font = Font(name='Times New Roman', size=10, bold=bold)
        cells.border = Border(bottom=thins, top=thins, left=thins, right=thins)
        cells.fill = PatternFill(fill_type='solid', fgColor='E6E6FA')
        column += 1
    row += 1
    column = 1
    #   TODO заполняем заголовок
    for it in title_sh:
        cells = ws.cell(row=row, column=column, value=it)
        cells.font = Font(name='Times New Roman', size=10, bold=bold)
        cells.border = Border(bottom=thins, top=thins)
        cells.fill = PatternFill(fill_type='solid', fgColor='FFFFE0')
        column += 1
    row += 1
    return row


def set_border(wse=None, cell_range='A1:F1'):
    thin = Side(border_style="thin", color="000000")
    for row in wse[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, bottom=thin)


def Read_Txt(command='1'):
    global section
    root = tkinter.Tk()
    root.title('Конвертор конфигурации С2000-М в список шлейфов')
    root.resizable(False, False)
    root.geometry('500x20')

    cfg_file = fd.askopenfilename(title='Открытие файла конфигурации С2000-М', filetypes=filetypes)
    #   TODO если файл не задан
    if cfg_file == '':
        sys.exit()
    #   TODO открываем файл конфигурации для чтения
    with open(cfg_file, 'r') as file_r:
        lines = file_r.readlines()
        #
    #   TODO заполняем именованный массив с разделами
    for line in lines:
        tes = line.replace('\n', '').split(', ')
        if tes.__len__() > 1 and tes[0].find('Раздел:') != -1 and tes[1].find('Описание:') != -1:
            te = int(str(tes[0]).replace(' ', '').replace('Раздел:', ''))
            section[te] = tes[1]
    tt = 1  # TODO row count
    for string_f in lines:
        #   TODO пропускаем пустые строки
        if string_f.find('\n') != -1:
            #   TODO    преобразуем строку в массив
            str_array = string_f.replace('\n', '').split(', ')
            #   TODO удаляем лишние прбелы из начала строки
            j = 0
            for h in str_array:
                if h.find('Описание:') != -1:
                    str_array[j] = h[h.find('Описание:'):]
                elif h.find('Шлейф:') != -1:
                    str_array[j] = h[h.find('Шлейф:'):]
                elif h.find('Раздел:') != -1:
                    str_array[j] = h[h.find('Раздел:'):]
                elif h.find('Время') != -1:
                    str_array[j] = h[h.find('Время'):]
                elif h.find('Выход:') != -1:
                    str_array[j] = h[h.find('Выход:'):]
                elif h.find('Реле:') != -1:
                    str_array[j] = h[h.find('Реле:'):]
                j += 1
                #   TODO
            if str_array[0].find('Конфигурация') != -1:
                tt = exls_w_titul(1, 1, str_array, True)
            if str_array[0].find('Версия:') != -1:
                tt = exls_w_titul(1, 1, str_array, True)
            #   TODO Заполняем строку с адресом и типом прибора
            if str_array.__len__() >= 3 and str_array[0].find('Адрес:') != -1:
                tt = exls_w_adr(tt, 1, str_array)
            #   TODO Заполняем строку с Шлейфом и описанием
            if str_array.__len__() >= 3 and str_array[0].find('Шлейф:') != -1:
                tt = exls_w_out(tt, 2, str_array)
            #   TODO Заполняем строку с Выходами
            if str_array.__len__() >= 2 and str_array[0].find('Выход:') != -1:
                tt = exls_w_out(tt, 2, str_array)
            #   TODO Заполняем строку с Реле
            if str_array.__len__() >= 2 and str_array[0].find('Реле:') != -1:
                tt = exls_w_out(tt, 2, str_array)
                '''
                #   TODO Заполняем ссценарии
                if test1.__len__() >= 3 and test1[0].find('Сценарий_упр:') != -1:
                    tt = Exls_w_sc(tt, 2, test1, True, tt+1)
                if test1.__len__() >= 1 and test1[0].find('Шаг:') != -1:
                    tt = Exls_w_sc(tt, 2, test1, True, tt)
#                if test1.__len__() >= 1 and test1[0].find('Состояния:') != -1:
#                    print(str(test1).replace('\'', ''))
#                    tt = Exls_w_sc(tt, 2, tuple(test1), True, tt)
                if test1.__len__() >= 1 and test1[0].find('Разделы: ') != -1:
                    tt = Exls_w_sc(row=tt, list_str=test1, bold=True, group_start=tt)
                if test1[0].find('Разрешающее условие') != -1 or test1[0].find('Запрещающее условие') != -1:
                    tt = Exls_w_sc(row=tt, list_str=test1, bold=True, group_start=tt)
                '''
    #   TODO Устанавливаем границы ячеек
    set_border(ws, 'A1:F' + str(tt - 1))
    #   TODO усли запущена с любым аргументом
    if command == '1':
        #   TODO ищем где нет описания и материмся
        for row in ws.iter_rows(min_row=1, min_col=6, max_col=6, max_row=tt - 1):
            for cell in row:
                if cell.value is None:
                    cell.fill = PatternFill(fill_type='solid', fgColor='FF6347')
                    cell.value = 'Где описание!!!???'
                    cell.font = Font(name='Times New Roman', size=10, bold=True, italic=True)
    #   TODO устанавливаем границы печати по 6 столбцу
    ws.col_breaks.brk = [Break(6)]
    #   TODO режим просмотра с границами печати
    ws.sheet_view.view = 'pageBreakPreview'
    #   TODO устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 25
    #   TODO Скрываем\показываем столбцы
    ws.column_dimensions['C'].hidden = False
    ws.column_dimensions['E'].hidden = True
    #   TODO устанавливаем тип бумаги и ориентацию листа
    ws.set_printer_settings(
        ws.PAPERSIZE_A4,
        ws.ORIENTATION_PORTRAIT
    )
    '''
    ws.protection.enable()
    ws.protection.password = '123456'
    '''
    #   TODO сохраняем результаты этого изврата
    wb.save('output.xlsx')
    #   TODO открываем эксель
    os.startfile('output.xlsx')


if __name__ == '__main__':
    if len(sys.argv) > 1:
        #        find_xlsx()
        Read_Txt()
    else:
        Read_Txt(command='2')
